import openpyxl

import sys
sys.path.append('E:/PycharmProjects/SeleniumTest/PythonWork/ArtApp/ArtAppPortProject')
from PortCommons import log_class
import os
'''
思路：需要封装一下功能：
1.excel单元格的读
2.excel单元格的写
3.excel的保存
'''

class OperateExcel:

    def __init__(self, path=None, sheet_name=None):
        if path:
            self.path = path
            self.sheet_name = sheet_name
        else:
            self.path = "../testcase/testcase_APP.xlsx"
            self.sheet_name = "V4.9.5"
            # log_file = r"E:\PycharmProjects\SeleniumTest\PythonWork\ArtApp\ArtAppPortProject\log\log.txt"
            # 获取当前项目地址
            BaseDir = os.path.dirname(os.path.dirname(__file__))
            # 获取log文件夹下的log
            log_file = os.path.join(BaseDir, 'log', 'log.txt')
            log = log_class.Logger(log_file)
            # 如果使用的是默认用例，在日志打印出来告警一下
            log.info("本次使用的是默认测试用例： " + str(self.path))
        self.data = self.get_data()

    # 获取表格对象
    def get_data(self):
        # self.wb = openpyxl.load_workbook(self.path)
        # self.ws = self.wb[self.sheet_name]
        # return self.ws
        try:
            self.wb = openpyxl.load_workbook(self.path)
            self.ws = self.wb[self.sheet_name]
        except Exception as e:
            print('获取worksheet失败', str(e))

    # 获取单元格
    def get_cell(self, row, column):
        try:
            self.cell = self.ws.cell(row=row, column=column)
            return self.cell
        except Exception as e:
            print('获取单元格失败', str(e))

    # 获取单元格的值
    def get_value(self, row, column):
        try:
            return self.ws.cell(row=row, column=column).value
        except Exception as e:
            print('获取数据失败', str(e))

    # 给单元格设值
    def set_value(self, row, column, value):
        try:
            self.ws.cell(row=row, column=column).value = value
        except Exception as e:
            print('写入数据失败', str(e))

    # 获取最大列数
    def get_max_column(self):
        return self.ws.max_column

    # 获取最大行数
    def get_max_row(self):
        return self.ws.max_row

    def save(self, path):
        self.wb.save(path)


if __name__ == '__main__':
    # test_path = "../testcase/testcase_APP.xlsx"
    # object1 = OperateExcel(test_path, 'V4.9.5')
    object1 = OperateExcel()
    max_column = object1.get_max_column()
    print("表格一共有：【" + str(max_column) + "】列")
    max_row = object1.get_max_row()
    print("表格一共有：【" + str(max_row) + "】行")
    print("-------------------------------------------")

    # 循环读取数据
    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            value = object1.get_value(i, j)
            if value:
                print(value, '\t', end='')
            else:
                print('\t\t', end='')
        print("")