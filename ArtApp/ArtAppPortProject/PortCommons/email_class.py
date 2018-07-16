# coding: utf-8
import smtplib
import email.mime.multipart
import email.mime.text
from email.mime.application import MIMEApplication
import openpyxl
import sys
import os
# 获取相对路径
sys.path.append(os.getcwd() + '\\PortCommons')
import log_class
import time

class SendMail:
    # 如果返回是Flase，那么就发邮件，else 通过
    # 整个方法放在报告生成之后，邮件发送之前
    def is_result_pass(self,test_path):
        # 捕获Excel文件打开失败异常
        try:
            wb = openpyxl.load_workbook(test_path)
        except Exception as e:
            print("打开Excel文件失败！", str(e))

        # 获取TestCase工作表名的对象
        s = wb.get_sheet_by_name("V4.9.5")
        # 定义一个空的数组
        enList = []
        '''
        进行测试结果判断，开始是用for循环取值然后判断发邮件，发现怎么都逻辑不对一直返回的都是True
        后来就新建一个空的数组，存入取出的所有测试结果，判断失败是否存在列表中
        '''
        # 遍历Excel内容
        for i in range(2, s.max_row+1):
            # 获取到“测试结果”的值
            ensure = s.cell(row=i, column=11).value
            # 添加到列表
            enList.append(ensure)

        log_file = "log.txt"
        log = log_class.Logger(log_file)

        # 日志记录系统当前时间
        now = time.strftime("%Y-%m-%d %H_%M_%S")
        log.info("本次接口自动化测试的时间为： " + now + " ")
        # 日志记录失败和成功的用例数
        failureNum = str(enList.count('失败'))
        log.info("一共失败用例个数为： " + failureNum + " ")
        passNum = str(enList.count('成功'))
        log.info("一共成功用例个数为： " + passNum + " ")

         # 判断是否发送邮件
        if '失败' in enList:
            # print("报告中存在错误或者失败的用例，需要发送邮件>>>")
            return False
        else:
            # print("报告用例均执行成功")
            return True
            # 获取Excel中“测试结果”，判断是否执行该case
            # if (ensure == "失败"):
            #     # print("报告中存在错误或者失败的用例，需要发送邮件>>>")
            #     return False
            # else:
            #     print("报告用例均执行成功")
            #     return True

    def send_mail(self, from1, to, title, content=None, type="plain", attach=None, pic=None):
        msg = email.mime.multipart.MIMEMultipart()
        # 生成包含多个邮件体的对象
        msg["from"] = from1
        # msg["to"] = to
        msg["to"] = ','.join(to)
        msg["subject"] = title

        # 邮件正文
        if (content != None):
            txt = email.mime.text.MIMEText(content, type)
            msg.attach(txt)

        # 文件附件
        if (attach != None):
            part = MIMEApplication(open(attach, "rb").read())
            part.add_header("Content-Disposition", "attachment", filename=attach)
            msg.attach(part)

        # jpg图片附件
        if (pic != None):
            jpgpart = MIMEApplication(open(pic, "rb").read())
            jpgpart.add_header("Conten-Dispositon", "attachment", filename=pic)  # 收到的图片后缀名变成了.bin格式，上网查了一下资料，没看懂，先不纠结了。
            msg.attach(jpgpart)

        # 发送邮件
        smtp = smtplib
        # 链接服务器，smtp地址+端口
        smtp = smtplib.SMTP("smtp.163.com", "25")

        # 设置为调试模式，console中显示
        # smtp.set_debuglevel(1)

        # 登录，用户名+密码
        smtp.login("13469975256@163.com", "l2875985")

        # 发送，from+to+内容
        smtp.sendmail(from1, to, str(msg))

        # 退出
        smtp.quit()
        print("发送邮件结束>>>>>")

if __name__ == '__main__':
    #设置变量并调用发送邮件
    from1 = "13469975256@163.com"
    to = "957949761@qq.com"
    # 标题
    titile = "接口测试报告"
    # 内容
    content = "发送邮件测试"
    # 附件
    attach = r"E:\PycharmProjects\SeleniumTest\PythonWork\ArtApp\ArtAppPortProject\reports\2018-06-19 11_07_41 report.xlsx"
    #pic="E:\\Python_study\\work\\excel\\test.jpg"

    #生成对象并调用方法
    mail = SendMail()
    res = mail.is_result_pass(attach)
    if res:
        print("用例全部通过无需发邮件，程序结束！")
    else:
        mail.send_mail(from1, to, titile, content, "html", attach)
        print("发送邮件结束>>>>>")