import re
import json
from openpyxl.styles import Font, colors, Alignment
import sys
import os
sys.path.append(os.getcwd() + '\\PortCommons')
import reg_class

''''
本class完成：
调用请求类发送请求并接受响应
根据响应结果进行判断是否正确
把执行结果保存到excel
'''

class InterfaceTest:
    def interface_test(self, num, interface_name, method, form, full_url, data, check_point, s, i, log):
        # 调用返回类
        rp = reg_class.MyRequests()
        # 判断传入的请求方法是否是get方法
        if (method.lower() == "get"):
            http_code, response, time = rp.get_requests(full_url, data)
            # print("get_requests=",response)
        elif (method.lower() == "post" and form.lower() == "json"):
            http_code, response, time = rp.post_json(full_url, data)
            # print("post_requests_json=",response)
        elif (method.lower() == "post" and form.lower() == "form"):
            http_code, response, time = rp.post_form(full_url, data)
            # print("post_requests_form=",response)
        else:
            log.error(num + " " + interface_name + ":请求方法错误，请确认[请求方法]字段是否正确！")
            return 400, method.lower()
            # print("请求方法错误，请确认[请求方法]字段是否正确！")

        # response=str(response)#re.search()是字符串的匹配，需要将返回的字典类型转换为字符串类型
        # print("response=",response)
        # print(response)
        # print(re.search("'error_code': 0",response).group())
        # print(check_point)
        # check_point=str(check_point)
        '''
        从excel获取到check_point="error_code": 0，error_code上面是双引号；而返回的response=
        {'error_code': 0, 'name': '小强软件测试疯狂讲义', 'reason': 'successed', 'price': 666}里面的'error_code': 0是单引号，导致
        re.search(check_point,response)时一直匹配不成功，所以要把从excel里获取到的error_code上的双引号替换成单引号，用re.sub()方法替换
        '''
        # check_point=re.sub('"',"'",check_point)
        # print(check_point)

        '''
        这里需要注意，咱们之前处理的都是把json类型转换成了python类型进行的
        处理完之后，为了显示效果，最后又转换成了json类型了，用的方法就是json.dumps()
        其中的参数ensure_ascii=False是为了解决中文编码
        '''
        response_json = json.dumps(response, ensure_ascii=False)
        # print("response_json=",response_json)
        if (http_code == 200):
            if (re.search(check_point, response_json)):
                # s=wb.get_sheet_by_name("TestCase")#获取TestCase工作表名的对象
                # print(i)#需要从run()传入当前循环i值
                s.cell(row=i, column=11).value = "成功"
                s.cell(row=i, column=12).value = response_json
                s.cell(row=i, column=13).value = time
                # 写入日志>>接口测试成功信息
                # 先注释，今后放开>>>>
                # log.info(">>编号=" + num + " " + "接口名称=" + interface_name + ",成功" + " " + "http状态码=" + str(http_code) \
                #          + " " + "响应时间=" + str(time) + "秒" + "\n" + "响应内容=" + response_json)
            else:
                myfont = Font(color=colors.RED)
                s.cell(row=i, column=11).value = "失败"
                # 失败的单元格设置成红色
                s.cell(row=i, column=11).font = myfont
                s.cell(row=i, column=12).value = response_json
                log.error(num + " " + interface_name + ",失败！, [ " + str(http_code) + " ], " + response_json)
                return 2001, response_json
        else:
            myfont = Font(color=colors.RED)
            s.cell(row=i, column=11).value = "失败"
            # 失败的单元格设置成红色
            s.cell(row=i, column=11).font = myfont
            s.cell(row=i, column=12).value = response_json
            log.error(num + " " + interface_name + ",失败！,[" + str(http_code) + " ], " + response_json)
        # 返回response响应结果，run_testcase里需要用来获取关联参数的值
        return response

        # 原来每次循环的时候都重新打开了一下excel，把之前写的又给清空了，应该把测试的sheet对象传过来了
        # wb.save("E:\\Python_study\\work\\project\\reports\\testcase_report.xlsx")
        # 这样循环写完之后，可以回到run()函数进行excel的保存