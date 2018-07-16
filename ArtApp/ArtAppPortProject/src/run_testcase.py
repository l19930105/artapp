import openpyxl
import sys
import os
# 获取相对路径
sys.path.append(os.getcwd() + '\\PortCommons')
import log_class
import read_init
sys.path.append(os.getcwd() + '\\src')
import interface_test
'''
本class完成：
一条条遍历excel里可执行的case
每次遍历调用interface_test.py的中方法进行请求
'''
class RunTestCase:
    def __init__(self,logfile):
        logfile = "log.txt"
        log = log_class.Logger(logfile)  # 引入日志类
        self.log = log
        read_ini = read_init.ReadIni()   # 引入获取配置文件类
        self.read_ini = read_ini

    def run(self, test_path, report_path):
        # 字典，为后续保存关联响应数据做准备
        correlationDict = {}

        # 捕获Excel文件打开失败异常
        try:
            wb = openpyxl.load_workbook(test_path)
        except Exception as e:
            print("打开Excel文件失败！", str(e))

        # 获取TestCase工作表名的对象
        s = wb.get_sheet_by_name("V4.9.5")
        # 获取执行的用例数，并写入日志
        amount = str(int(s.cell(s.max_row, column=1).value))
        self.log.info("一共运行了 " + amount + "个用例")

        # 遍历Excel内容
        for i in range(2, s.max_row+1):
            # 获取Excel中“是否运行”，判断是否执行该case
            if (s.cell(row=i, column=10).value != "Yes"):
                continue

            # 给地址前缀赋值为测试环境的地址 --->具体环境配置请看config.ini文件
            # oAddress-->生产环境    tAddress-->测试环境
            address = self.read_ini.get_value("tAddress","testAress")
            s.cell(row=i, column=3).value = address
            # print(">>>>当前测试地址为： " + s.cell(row=i, column=3).value)

            # 获取excel里编号得到的int类型，int类型需要转换成string类型才能做replace
            num = str(int(s.cell(row=i, column=1).value)).replace("\n"," ").replace("\r"," ")
            # 获取excel里“接口名称”
            interface_name = (s.cell(row=i, column=2).value).replace("\n", "").replace("\r", "")
            # 获取excel里“地址前缀”
            # url = (s.cell(row=i, column=3).value).replace("\n", "").replace("\r", "")
            url = (s.cell(row=i, column=3).value).replace("\n", "").replace("\r", "")
            # 获取excel里“请求地址”
            path = (s.cell(row=i, column=4).value).replace("\n", "").replace("\r", "")
            # 获取excel里“请求方法”
            method = (s.cell(row=i, column=5).value).replace("\n", "").replace("\r", "")
            # 获取excel里“请求格式”
            form = (s.cell(row=i, column=6).value).replace("\n", "").replace("\r", "")
            # 获取excel里“请求数据”
            data = (s.cell(row=i, column=7).value).replace("\n", "").replace("\r", "")

            # 如果接口参数很长时，可以把参数放到文本文件里，excel里对应的请求数据写入文本文件的路径，然后读取文本文件的内容
            # 这里注意txt的编码必须是utf-8无bom格式的，可以用Notepad++
            if(os.path.exists(data)):
                fopen = open(data, encoding="utf-8")
                data = fopen.readline()
                fopen.close()

            '''
            写完关联参数的拆分后，开始把关联参数的替换也写在了下面，执行报错，看报错信息然后分析出是有问题的，放在这下面替换的是当前请求，然后需要替换的参数的应该是下一个循环的请求，所以应该放到发送请求的代码前
            '''
            # 在data中查找是否存在需要替换的数据
            for keyword in correlationDict:
                if data.find(keyword) > 0:
                    data = data.replace(keyword, str(correlationDict[keyword]))
                    # print("data=", data)

            # #注意从表里取出来的都是str的类型，要用eval转成字典啊
            data = eval(data)

            # 获取excel里“检查点”
            check_point = s.cell(row=i, column=8).value.replace("\n", "").replace("\r", "")
            # 获取excel里“关联参数”
            correlation = s.cell(row=i, column=9).value
            # 调用InterfaceTest类
            it = interface_test.InterfaceTest()
            # 拼接完整url
            full_url = url + path

            # 需要从interface_test()方法里把响应结果返回，用来取关联参数的值
            response = it.interface_test(num, interface_name, method, form, full_url, data, check_point, s, i, self.log)

            # 对获取到的关联参数进行拆分，获取参数对应的响应数据
            if(correlation != None):
                # 去掉多余的回车换行，如果有多个关联参数则用分号隔开
                correlation = correlation.replace("\n", "").replace("\r", "").split(';')
                for j in range(len(correlation)):
                    # 根据=把关联数据拆分
                    param = correlation[j].split('=')

                    # 一般等号两边是两个元素，所以判断等于2
                    if len(param) == 2:
                        '''
                        我们关注第二个元素，需要把他取出来（因为他就是我们需要的响应数据）
                        从下标1开始到末尾
                        那为什么后面还有split呢？看下这样的数据
                        ${date}=[result][yangli]
                        '''
                        # [1:-1]代表取出“[checkstatus]”的 checkstatus
                        for key in param[1][1:-1].split(']['):
                            # print("key=", key)
                            # 根据第二个元素把对应的响应数据拿到
                            temp = response[key]
                            # print("temp=", temp)
                            # 因为中间处理的时候数据会有变化，所以在给一个新的值存储
                            response = temp
                            # print("response=", response)

                        # 关联到的响应放到字典里，方便后续去遍历替换参数
                        correlationDict[param[0]] = response
                        # print("correlationDict[param[0]]=", response)
                    else:
                        # print("error")
                        self.log.error("请检查excel中关联参数的配置")
        # 接口测试的结果存在结果用例里
        wb.save(report_path)